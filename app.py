import io
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from thefuzz import fuzz

st.set_page_config(
    page_title="Data Name Fuzzy Merger",
    page_icon="👥",
    layout="wide",
)

st.title("👥 Data Name Fuzzy Merger & Painter")
st.write(
    "Unggah beberapa file Excel. Aplikasi akan mencocokkan data berdasarkan **No. Rekening**, "
    "menganalisis kemiripan isi **Nama Karyawan**, tetap menggunakan nama dari file acuan pertama, dan mewarnai sel nama tersebut sesuai tingkat kemiripannya."
)

uploaded_files = st.file_uploader(
    "Upload Semua File Excel Anda (Minimal 2 file)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if uploaded_files:
    if len(uploaded_files) < 2:
        st.warning("⚠️ Silahkan unggah minimal 2 file Excel untuk digabungkan.")
    else:
        st.info(
            f"💡 File acuan utama: **{uploaded_files.name}**. Seluruh nama karyawan akan mengacu pada file ini."
        )

        if st.button("🚀 Mulai Gabungkan & Analisis Data", use_container_width=True):
            try:
                # 1. BACA FILE ACUAN UTAMA
                df_master = pd.read_excel(uploaded_files)
                df_master.columns = [str(c).strip() for c in df_master.columns]
                
                # Validasi kolom wajib di file pertama
                if "Nama Karyawan" not in df_master.columns or "No. Rekening" not in df_master.columns:
                    st.error("File pertama wajib memiliki kolom 'Nama Karyawan' dan 'No. Rekening'!")
                    st.stop()
                
                df_master["No. Rekening"] = df_master["No. Rekening"].astype(str).str.strip()
                
                # Buat DataFrame penanda warna khusus untuk kolom 'Nama Karyawan'
                # Default 'H' (Hijau) untuk file pertama karena 100% mirip dirinya sendiri
                warna_nama_list = ["H"] * len(df_master)

                # 2. PROSES FILE TAMBAHAN SATU PER SATU
                for file_tambahan in uploaded_files[1:]:
                    df_temp = pd.read_excel(file_tambahan)
                    df_temp.columns = [str(c).strip() for c in df_temp.columns]
                    
                    if "Nama Karyawan" not in df_temp.columns or "No. Rekening" not in df_temp.columns:
                        st.error(f"File '{file_tambahan.name}' wajib memiliki kolom 'Nama Karyawan' and 'No. Rekening'!")
                        st.stop()
                        
                    df_temp["No. Rekening"] = df_temp["No. Rekening"].astype(str).str.strip()

                    # Buat dictionary file tambahan berdasarkan No. Rekening untuk pencarian cepat
                    # Struktur: { 'no_rek': row_data_dict }
                    dict_temp = {}
                    for _, row in df_temp.iterrows():
                        dict_temp[str(row["No. Rekening"])] = row.to_dict()

                    # Siapkan list untuk menampung baris data baru dari file tambahan ini
                    baris_baru_list = []

                    for _, row_master in df_master.iterrows():
                        norek = str(row_master["No. Rekening"])
                        
                        if norek in dict_temp:
                            row_temp = dict_temp[norek]
                            nama1 = str(row_master["Nama Karyawan"]).strip()
                            nama2 = str(row_temp["Nama Karyawan"]).strip()
                            
                            # HITUNG FUZZY MATCH PADA DATA NAMA
                            skor = fuzz.ratio(nama1.lower(), nama2.lower())
                            
                            # Tentukan flag warna berdasarkan skor kemiripan teks data
                            if skor > 80:
                                flag_warna = "H"
                            elif 30 <= skor <= 80:
                                flag_warna = "K"
                            else:
                                flag_warna = "M"
                                
                            # Gabungkan kolom baru dari file tambahan ke baris ini (jika ada kolom baru)
                            for col in row_temp.keys():
                                if col not in df_master.columns:
                                    df_master[col] = ""  # Buat kolom baru di master jika belum ada
                                    
                            # Update nilai kolom tambahan di data master saat ini
                            for col, val in row_temp.items():
                                if col != "Nama Karyawan" and col != "No. Rekening":
                                    df_master.at[_, col] = val
                                    
                            warna_nama_list[_] = flag_warna
                        else:
                            # Jika nomor rekening tidak ditemukan di file tambahan, beri warna merah pada baris tersebut
                            if warna_nama_list[_] not in ["H", "K"]:
                                warna_nama_list[_] = "M"

                # Isi nilai kosong (NaN) dengan string kosong agar rapi
                df_master = df_master.fillna("")

                # 3. PRATINJAU TABEL DI WEBSITE
                st.subheader("👀 Pratinjau Tabel Hasil Gabungan")
                st.write("Warna di bawah ini diterapkan khusus pada kolom **Nama Karyawan** berdasarkan akurasi kemiripan teks datanya.")

                def style_nama_column(df_input):
                    df_style = pd.DataFrame("", index=df_input.index, columns=df_input.columns)
                    for r in range(len(df_input)):
                        flag = warna_nama_list[r]
                        if flag == "H":
                            df_style.loc[df_input.index[r], "Nama Karyawan"] = "background-color: #C6EFCE; color: #006100"
                        elif flag == "K":
                            df_style.loc[df_input.index[r], "Nama Karyawan"] = "background-color: #FFEB9C; color: #9C0006"
                        elif flag == "M":
                            df_style.loc[df_input.index[r], "Nama Karyawan"] = "background-color: #FFC7CE; color: #9C0006"
                    return df_style

                st.dataframe(df_master.style.apply(style_nama_column, axis=None), use_container_width=True)

                # 4. PROSES PEWARNAAN CELL EXCEL ASLI (OPENPYXL)
                buffer = io.BytesIO()
                df_master.to_excel(buffer, index=False)
                buffer.seek(0)

                wb = load_workbook(buffer)
                ws = wb.active

                warna_hijau = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                warna_kuning = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                warna_merah = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # Cari indeks kolom 'Nama Karyawan' di file excel hasil
                idx_nama = next(col for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value == "Nama Karyawan")

                # Warnai cell nama karyawan berdasarkan list warna_nama_list
                for r_idx in range(2, ws.max_row + 1):  # Mulai baris 2 (data)
                    flag_val = warna_nama_list[r_idx - 2]
                    cell = ws.cell(row=r_idx, column=idx_nama)

                    if flag_val == "H":
                        cell.fill = warna_hijau
                    elif flag_val == "K":
                        cell.fill = warna_kuning
                    elif flag_val == "M":
                        cell.fill = warna_merah

                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)

                st.success("🎉 Analisis kemiripan isi data Nama Karyawan selesai!")

                # 5. TOMBOL DOWNLOAD
                st.download_button(
                    label="📥 Unduh File Gabungan (.xlsx)",
                    data=output_buffer,
                    file_name="gabungan_data_nama_fuzzy.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Terjadi kesalahan saat memproses data: {e}")
                df_master = pd.read_excel(file_acuan)

                # Format nama kolom file acuan agar bersih
                df_master.columns = [str(c).strip() for c in df_master.columns]
                kolom_acuan_asli = list(df_master.columns)

                # Tambahkan kolom penanda warna: 'H' = Hijau, 'M' = Merah
                df_master["_row_color_flag"] = "H"

                # 3. PROSES FILE KEDUA DAN SETERUSNYA
                for file_tambahan in uploaded_files[1:]:
                    df_temp = pd.read_excel(file_tambahan)
                    df_temp.columns = [str(c).strip() for c in df_temp.columns]

                    # Hitung kemiripan (similarity) set kolom
                    kolom_master_set = set(kolom_acuan_asli)
                    kolom_temp_set = set(df_temp.columns)
                    kolom_sama = kolom_master_set.intersection(kolom_temp_set)

                    total_kolom_gabungan = len(
                        kolom_master_set.union(kolom_temp_set)
                    )
                    skor_kemiripan_kolom = (
                        (len(kolom_sama) / total_kolom_gabungan) * 100
                        if total_kolom_gabungan > 0
                        else 0
                    )

                    # Tentukan penanda warna baris
                    if skor_kemiripan_kolom >= 80:
                        df_temp["_row_color_flag"] = "H"
                    else:
                        df_temp["_row_color_flag"] = "M"

                    # Gabungkan data
                    df_master = pd.concat([df_master, df_temp], ignore_index=True)

                # Isi nilai kosong (NaN) dengan string kosong
                df_master = df_master.fillna("")

                # Susun ulang urutan kolom
                semua_kolom = list(df_master.columns)
                semua_kolom.remove("_row_color_flag")
                kolom_baru = [c for c in semua_kolom if c not in kolom_acuan_asli]
                urutan_kolom_final = (
                    kolom_acuan_asli + kolom_baru + ["_row_color_flag"]
                )
                df_master = df_master[urutan_kolom_final]

                # 4. FITUR PRATINJAU TABEL (PREVIEW DATA) DI WEB
                st.subheader("👀 Pratinjau Tabel Hasil Gabungan")
                st.write(
                    "Berikut adalah tampilan sementara data Anda. Baris dengan latar belakang "
                    "hijau/merah menandakan tingkat kecocokan kolom."
                )

                # Fungsi styling untuk pratinjau tabel Streamlit
                def style_row(row):
                    # Ambil flag warna di kolom terakhir
                    flag = row["_row_color_flag"]
                    if flag == "H":
                        return ["background-color: #C6EFCE; color: #006100"] * len(
                            row
                        )
                    elif flag == "M":
                        return ["background-color: #FFC7CE; color: #9C0006"] * len(
                            row
                        )
                    return [""] * len(row)

                # Tampilkan dataframe dengan style warna (kolom flag disembunyikan di preview)
                df_preview = df_master.copy()
                st.dataframe(
                    df_preview.style.apply(style_row, axis=1).hide(
                        axis="columns", subset=["_row_color_flag"]
                    ),
                    use_container_width=True,
                )

                # 5. LOGIKA PEWARNAAN FILE EXCEL ASLI (OPENPYXL)
                buffer = io.BytesIO()
                df_master.to_excel(buffer, index=False)
                buffer.seek(0)

                wb = load_workbook(buffer)
                ws = wb.active

                warna_hijau = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
                warna_merah = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )

                idx_flag = ws.max_column

                for row in range(2, ws.max_row + 1):
                    flag_val = ws.cell(row=row, column=idx_flag).value
                    if flag_val == "H":
                        for col in range(1, idx_flag):
                            ws.cell(row=row, column=col).fill = warna_hijau
                    elif flag_val == "M":
                        for col in range(1, idx_flag):
                            ws.cell(row=row, column=col).fill = warna_merah

                # Hapus kolom bantuan flag agar file hasil bersih
                ws.delete_cols(idx_flag)

                output_buffer = io.BytesIO()
                wb.save(output_buffer)
                output_buffer.seek(0)

                st.success(
                    "🎉 Analisis selesai! Struktur warna di atas akan sama persis dengan file Excel yang diunduh."
                )

                # 6. TOMBOL DOWNLOAD HASIL AKHIR
                st.download_button(
                    label="📥 Unduh File Gabungan (.xlsx)",
                    data=output_buffer,
                    file_name="gabungan_excel_with_preview.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Terjadi kesalahan saat memproses file: {e}")
